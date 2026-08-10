"""Parse Kiwoom/KIS API workbooks into structured specs.

The vendor workbooks are perfectly regular: sheet 0 is a catalog, and every other
sheet documents exactly one API with the same layout -- a key/value info block,
then ``Request``/``Response`` field tables, then examples. That regularity means
extraction is deterministic; no model is involved. See ``spec_store`` for how the
parsed specs are served to the agent.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path

import openpyxl


class Market(StrEnum):
    """Kiwoom exposes domestic and US equities as separate API surfaces.

    They share a host but nothing else: different URL trees, different realtime
    websocket endpoints, and separate trading entitlements. Never route an id
    from one market through a client for the other -- hence a real type rather
    than a string filter.
    """

    KR = "KR"
    US = "US"
    AUTH = "AUTH"
    COMMON = "COMMON"


def market_of(url: str) -> Market:
    if url.startswith("/api/us"):
        return Market.US
    if url.startswith("/api/dostk"):
        return Market.KR
    if url.startswith("/oauth2"):
        return Market.AUTH
    return Market.COMMON


# Label -> ApiSpec attribute, as they appear in column A of the info block.
_INFO_LABELS = {
    "메뉴 위치": "menu_path",
    "API 명": "name",
    "API ID": "api_id",
    "Method": "method",
    "운영 도메인": "prod_domain",
    "모의투자 도메인": "mock_domain",
    "URL": "url",
    "Format": "format",
    "Content-Type": "content_type",
}

_SECTION_REQUEST = "Request"
_SECTION_RESPONSE = "Response"
_SECTION_REQUEST_EXAMPLE = "Request Example"
_SECTION_RESPONSE_EXAMPLE = "Response Example"
_SECTION_OVERVIEW = "개요"
_TABLE_HEADER = "구분"
_SECTIONS = {
    _SECTION_REQUEST,
    _SECTION_RESPONSE,
    _SECTION_REQUEST_EXAMPLE,
    _SECTION_RESPONSE_EXAMPLE,
    _SECTION_OVERVIEW,
    "기본정보",
    "API 정보",
}

# Nested list members are marked by a leading dash on the element name.
_NESTED = re.compile(r"^-\s*")


@dataclass(slots=True)
class Field:
    """One row of a Request/Response field table."""

    element: str
    korean_name: str = ""
    type: str = ""
    required: bool = False
    length: str = ""
    description: str = ""
    section: str = "Body"  # Header | Body
    parent: str | None = None  # set for `- foo` rows nested under a list element

    @property
    def is_container(self) -> bool:
        """True when other fields declare this element as their parent."""
        return self.type == "" and self.length == "" and self.section == "Body"


@dataclass(slots=True)
class ApiSpec:
    api_id: str
    name: str = ""
    menu_path: str = ""
    method: str = "POST"
    url: str = ""
    prod_domain: str = ""
    mock_domain: str = ""
    format: str = "JSON"
    content_type: str = ""
    overview: str = ""
    request: list[Field] = field(default_factory=list)
    response: list[Field] = field(default_factory=list)
    request_example: str = ""
    response_example: str = ""

    # -- convenience views used by the client and by prompt rendering ---------

    def body_fields(self, side: str = "request") -> list[Field]:
        fields = self.request if side == "request" else self.response
        return [f for f in fields if f.section == "Body"]

    def required_body(self) -> list[str]:
        return [f.element for f in self.body_fields() if f.required and f.parent is None]

    def known_body(self) -> set[str]:
        return {f.element for f in self.body_fields() if f.parent is None}

    def domain(self, *, testnet: bool) -> str:
        return self.mock_domain if testnet else self.prod_domain

    @property
    def is_websocket(self) -> bool:
        """Realtime subscriptions advertise a `wss://` domain, not `https://`."""
        return self.prod_domain.startswith("wss://")

    @property
    def market(self) -> Market:
        return market_of(self.url)

    @property
    def category(self) -> str:
        """Top-level menu category, e.g. `국내주식` from `국내주식 > 주문 > ...`."""
        return self.menu_path.split(">")[0].strip() if self.menu_path else ""


def _cells(row: tuple, width: int = 8) -> list[str]:
    out = []
    for i in range(width):
        v = row[i] if i < len(row) else None
        out.append("" if v is None else str(v).strip())
    return out


def _first_value(cells: list[str], start: int = 1) -> str:
    """Info-block values sit a variable number of columns right of the label."""
    for c in cells[start:]:
        if c:
            return c
    return ""


def _parse_table(rows: list[list[str]], start: int) -> tuple[list[Field], int]:
    """Read a field table beginning at the `구분` header row at index `start`.

    Returns the fields and the index of the first row after the table.
    """
    fields: list[Field] = []
    section = "Body"
    last_top: str | None = None
    i = start + 1
    while i < len(rows):
        cells = rows[i]
        head, element = cells[0], cells[1]
        # A new section marker (or another table header) ends this table.
        if head in _SECTIONS or head == _TABLE_HEADER:
            break
        if head in ("Header", "Body"):
            section = head
            last_top = None
        if not element:
            # Blank element with no section marker: end of table body.
            if not any(cells):
                break
            i += 1
            continue

        parent = None
        name = element
        if _NESTED.match(element):
            name = _NESTED.sub("", element).strip()
            parent = last_top

        f = Field(
            element=name,
            korean_name=cells[2],
            type=cells[3],
            required=cells[4].upper().startswith("Y"),
            length=cells[5],
            description=cells[6],
            section=section,
            parent=parent,
        )
        fields.append(f)
        if parent is None and section == "Body":
            last_top = name
        i += 1
    return fields, i


def parse_sheet(ws) -> ApiSpec:
    """Parse one API sheet into an :class:`ApiSpec`."""
    rows = [_cells(r) for r in ws.iter_rows(values_only=True)]
    spec = ApiSpec(api_id="")
    overview_parts: list[str] = []

    i = 0
    while i < len(rows):
        cells = rows[i]
        head = cells[0]

        if head in _INFO_LABELS:
            setattr(spec, _INFO_LABELS[head], _first_value(cells))
            i += 1
            continue

        if head == _SECTION_OVERVIEW:
            i += 1
            # Free text until the next recognised section marker.
            while i < len(rows) and rows[i][0] not in _SECTIONS and rows[i][0] not in _INFO_LABELS:
                if rows[i][0]:
                    overview_parts.append(rows[i][0])
                i += 1
            continue

        if head in (_SECTION_REQUEST, _SECTION_RESPONSE):
            side = head
            # The table header row follows, possibly after blank rows.
            j = i + 1
            while j < len(rows) and rows[j][0] != _TABLE_HEADER:
                if rows[j][0] in _SECTIONS:
                    break
                j += 1
            if j < len(rows) and rows[j][0] == _TABLE_HEADER:
                parsed, nxt = _parse_table(rows, j)
                if side == _SECTION_REQUEST:
                    spec.request = parsed
                else:
                    spec.response = parsed
                i = nxt
                continue
            i = j
            continue

        if head in (_SECTION_REQUEST_EXAMPLE, _SECTION_RESPONSE_EXAMPLE):
            body: list[str] = []
            j = i + 1
            while j < len(rows) and rows[j][0] not in _SECTIONS:
                if rows[j][0]:
                    body.append(rows[j][0])
                j += 1
            text = "\n".join(body)
            if head == _SECTION_REQUEST_EXAMPLE:
                spec.request_example = text
            else:
                spec.response_example = text
            i = j
            continue

        i += 1

    spec.overview = " ".join(overview_parts).strip()
    return spec


@dataclass(slots=True)
class CatalogEntry:
    api_id: str
    name: str
    category: str = ""
    subcategory: str = ""
    url: str = ""

    @property
    def market(self) -> Market:
        return market_of(self.url)


def parse_catalog(ws) -> list[CatalogEntry]:
    """Parse the index sheet (`API 리스트`) into catalog entries."""
    entries: list[CatalogEntry] = []
    seen_header = False
    for row in ws.iter_rows(values_only=True):
        cells = _cells(row)
        if not seen_header:
            if cells[1] == "API ID":
                seen_header = True
            continue
        api_id = cells[1]
        if not api_id:
            continue
        entries.append(
            CatalogEntry(
                api_id=api_id,
                name=cells[2],
                category=cells[3],
                subcategory=cells[4],
                url=cells[5],
            )
        )
    return entries


def parse_workbook(path: str | Path) -> tuple[list[CatalogEntry], dict[str, ApiSpec]]:
    """Parse a full vendor workbook.

    Returns the catalog and a mapping of API ID -> spec. Specs are keyed by the
    `API ID` in each sheet's info block, not by sheet name: REST sheets are named
    by bare id (`ka10001`) while realtime sheets are named `이름(id)` (`주문체결(00)`),
    and one sheet name even substitutes `|` for `/`. Reference sheets with no API
    ID (e.g. the error-code table) fall back to their sheet name.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        catalog = parse_catalog(wb[wb.sheetnames[0]])
        specs: dict[str, ApiSpec] = {}
        for name in wb.sheetnames[1:]:
            spec = parse_sheet(wb[name])
            specs[spec.api_id or name] = spec
        return catalog, specs
    finally:
        wb.close()
